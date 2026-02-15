# coding: utf-8
from sklearn import linear_model
import matplotlib.pyplot as plt
from sklearn.ensemble import RandomForestRegressor
from sklearn.preprocessing import MinMaxScaler
from sklearn.preprocessing import StandardScaler
import numpy as np
import pandas as pd
from fndata import Extraction_module
from fndata import ResultUpload_module
from sklearn.metrics import confusion_matrix
from openpyxl import load_workbook
import tensorflow as tf
from tensorflow import keras
import math
from datetime import datetime

class DataLoad(object):
    @staticmethod
    def fndata_dataframe(item_name, target_column, merge=True):
        #addr = "165.246.45.192"
        #host = "sunggil"
        #pwd = "ihq123!@#"
        #schema = "financedb"

        addr, host, pwd, schema = '165.246.196.138', 'vf-user', '!@value1419', 'financedb'

        conn = Extraction_module(addr, host, pwd, schema)

        item = conn.item_data_load(item_name, merge=True)
        market = conn.exogenous_data_load('market', merge=True)
        economy = conn.exogenous_data_load('economy', merge=True)

        if merge == False:
            return item, market, economy

        data_dict = dict()
        data_dict['item'] = item
        data_dict['market'] = market
        data_dict['economy'] = economy

        data = pd.concat([item, market], axis=1)
        data = pd.concat([data, economy], axis=1)
        data.columns = pd.io.parsers.ParserBase({'names': data.columns})._maybe_dedup_names(data.columns)  # 중복컬럼명 처리
        data = data.dropna(subset=[target_column])  # 타겟 컬럼이 nan이면 행 전체 제거
        data = data.reset_index(drop=True)

        return data


class FeatureSelection(object):
    @staticmethod
    def lasso_fit(X, y, alpha=1.0, model_return=False):
        model = linear_model.Lasso(alpha=alpha)  # 모델 생성
        model.fit(X, y)  # 모델 학습
        print("#success Lasso model training")
        feature_score = pd.DataFrame({'feature': X.columns, 'score': abs(model.coef_)})  # 변수 중요도(weight 절대값) 추출
        feature_sorted_score = feature_score.sort_values(by=['score'], ascending=False)  # score 정렬
        if model_return == True:
            return model, feature_sorted_score
        return feature_sorted_score

    @staticmethod
    def ridge_fit(X, y, alpha=1.0, model_return=False):
        model = linear_model.Ridge(alpha=alpha)
        model.fit(X, y)
        print("#success Ridge model training")
        feature_score = pd.DataFrame({'feature': X.columns, 'score': abs(model.coef_)})
        feature_sorted_score = feature_score.sort_values(by=['score'], ascending=False)
        if model_return == True:
            return model, feature_sorted_score
        return feature_sorted_score

    @staticmethod
    def randomforest_fit(X, y, n_estimators=500, n_jobs=-1, min_samples_leaf=50, model_return=False):
        model = RandomForestRegressor(n_estimators=n_estimators, n_jobs=n_jobs, min_samples_leaf=min_samples_leaf)
        model.fit(X, y)
        print("#success RF model training")
        feature_score = pd.DataFrame({'feature': X.columns, 'score': model.feature_importances_})
        feature_sorted_score = feature_score.sort_values(by=['score'], ascending=False)
        if model_return == True:
            return model, feature_sorted_score
        return feature_sorted_score

class CreateInputOutput(object):
    @staticmethod
    def input_2d_data(data, n_timestep, future_day, time_interval=1):
        data = data[:-future_day]
        input_2d_data = np.zeros((data.shape[0] + 1 - n_timestep, math.ceil(n_timestep / time_interval), data.shape[1]),
                                 dtype=np.float32)  # (size, n_timestep, n_features)

        for i in range(len(input_2d_data)):
            input_2d_data[i] = data[i:n_timestep + i:time_interval]

        return input_2d_data

    @staticmethod
    def many_to_many_output(target, n_timestep, time_interval=1):
        if type(target) != np.ndarray:
            target = np.array(target)

        output = np.zeros((target.shape[0] + 1 - n_timestep, math.ceil(n_timestep / time_interval), 1), dtype=np.float32)
        target = target.reshape(-1, 1)

        if time_interval == 1:
            for i in range(len(output)):
                output[i] = target[i:n_timestep + i:time_interval]

        else:
            for i in range(len(output)):
                temp = target[n_timestep + i - 1:i:-time_interval]
                output[i] = temp[::-1]

        return output

    @staticmethod
    def many_to_one_output(target, n_timestep):
        if type(target) != np.ndarray:
            target = np.array(target)
        target = target[n_timestep - 1:]  # 차원 추가
        output = np.expand_dims(target, axis=-1)
        return output

    # ==================== Transformer용 신규 메서드들 ====================
    
    @staticmethod
    def transformer_input_data(data, n_timestep, future_day, stride=1):
        """
        Transformer용 입력 데이터 생성
        time_interval 없이 연속된 시계열 데이터 사용
        
        Args:
            data: 원본 데이터 (2D array: [날짜 수, 특징 수])
            n_timestep: 시계열 길이 (예: 60일)
            future_day: 미래 예측 일수 (예: 20일)
            stride: 윈도우 이동 간격 (기본값 1 = 하루씩 이동)
        
        Returns:
            3D array: (샘플 수, n_timestep, 특징 수)
            
        예시:
            data.shape = (1000, 300)  # 1000일, 300개 특징
            n_timestep = 60
            future_day = 20
            stride = 1
            → output.shape = (920, 60, 300)  # 1000 - 60 - 20 + 1 = 921 샘플
        """
        # future_day만큼 앞의 데이터 제거 (예측할 데이터가 있어야 하므로)
        data = data[:-future_day]
        
        # 생성할 샘플 개수 계산
        n_samples = (data.shape[0] - n_timestep) // stride + 1
        
        # 출력 배열 초기화
        input_data = np.zeros((n_samples, n_timestep, data.shape[1]), dtype=np.float32)
        
        # 슬라이딩 윈도우로 데이터 생성
        for i in range(n_samples):
            start_idx = i * stride
            end_idx = start_idx + n_timestep
            input_data[i] = data[start_idx:end_idx]
        
        return input_data
    
    
    @staticmethod
    def transformer_output_sequence(target, n_timestep, future_day, stride=1, output_type='many_to_many'):
        """
        Transformer용 출력 데이터 생성
        
        Args:
            target: 타겟 값 (1D array: 수익률 등)
            n_timestep: 입력 시계열 길이
            future_day: 미래 예측 일수
            stride: 윈도우 이동 간격
            output_type: 'many_to_many' 또는 'many_to_one'
        
        Returns:
            many_to_many: (샘플 수, n_timestep, 1) - 각 시간스텝마다 예측값
            many_to_one: (샘플 수, 1) - 마지막 예측값만
            
        예시:
            target = [0.5, -0.3, 1.2, ...] (1000개)
            n_timestep = 60
            output_type = 'many_to_many'
            → output.shape = (921, 60, 1)
        """
        if type(target) != np.ndarray:
            target = np.array(target)
        
        n_samples = (len(target) - n_timestep - future_day + 1) // stride
        
        if output_type == 'many_to_many':
            # 각 시간스텝마다 예측값 생성 (LSTM과 동일)
            output = np.zeros((n_samples, n_timestep, 1), dtype=np.float32)
            
            for i in range(n_samples):
                start_idx = i * stride
                end_idx = start_idx + n_timestep
                output[i] = target[start_idx:end_idx].reshape(-1, 1)
            
        elif output_type == 'many_to_one':
            # 마지막 예측값만 생성
            output = np.zeros((n_samples, 1), dtype=np.float32)
            
            for i in range(n_samples):
                target_idx = i * stride + n_timestep - 1
                output[i] = target[target_idx]
        
        else:
            raise ValueError("output_type must be 'many_to_many' or 'many_to_one'")
        
        return output
    
    
    @staticmethod
    def transformer_future_target(target, n_timestep, future_day, stride=1):
        """
        미래 예측값을 타겟으로 하는 출력 생성
        현재 시점에서 future_day 이후의 값을 예측
        
        Args:
            target: 타겟 값 (수익률 등)
            n_timestep: 입력 시계열 길이
            future_day: 예측할 미래 일수
            stride: 윈도우 이동 간격
        
        Returns:
            (샘플 수, 1) - future_day 이후의 값
            
        예시:
            입력 윈도우: [Day 1 ~ Day 60]
            출력: Day 80의 수익률 (future_day=20)
        """
        if type(target) != np.ndarray:
            target = np.array(target)
        
        n_samples = (len(target) - n_timestep - future_day + 1) // stride
        output = np.zeros((n_samples, 1), dtype=np.float32)
        
        for i in range(n_samples):
            # 입력 윈도우의 마지막 시점 + future_day
            future_idx = i * stride + n_timestep + future_day - 1
            output[i] = target[future_idx]
        
        return output
    
    
    @staticmethod
    def transformer_multi_horizon_output(target, n_timestep, future_days_list, stride=1):
        """
        여러 미래 시점을 동시에 예측하는 출력 생성
        
        Args:
            target: 타겟 값
            n_timestep: 입력 시계열 길이
            future_days_list: 예측할 미래 일수 리스트 (예: [5, 10, 20])
            stride: 윈도우 이동 간격
        
        Returns:
            (샘플 수, len(future_days_list)) - 여러 미래 시점의 값
            
        예시:
            future_days_list = [5, 10, 20]
            → 5일 후, 10일 후, 20일 후를 동시에 예측
        """
        if type(target) != np.ndarray:
            target = np.array(target)
        
        max_future = max(future_days_list)
        n_samples = (len(target) - n_timestep - max_future + 1) // stride
        output = np.zeros((n_samples, len(future_days_list)), dtype=np.float32)
        
        for i in range(n_samples):
            for j, future_day in enumerate(future_days_list):
                future_idx = i * stride + n_timestep + future_day - 1
                output[i, j] = target[future_idx]
        
        return output
    
    
    @staticmethod
    def transformer_attention_mask(n_timestep, mask_future=True):
        """
        Transformer용 어텐션 마스크 생성
        
        Args:
            n_timestep: 시계열 길이
            mask_future: True면 미래 시점 마스킹 (causal mask)
        
        Returns:
            마스크 배열: (n_timestep, n_timestep)
            
        예시:
            mask_future=True일 때:
            [[1, 0, 0, 0],    # 1번째 시점은 자기 자신만 볼 수 있음
             [1, 1, 0, 0],    # 2번째 시점은 1~2번만 볼 수 있음
             [1, 1, 1, 0],    # 3번째 시점은 1~3번만 볼 수 있음
             [1, 1, 1, 1]]    # 4번째 시점은 1~4번 모두 볼 수 있음
        """
        if mask_future:
            # Causal mask (미래를 볼 수 없음)
            mask = np.tril(np.ones((n_timestep, n_timestep)))
        else:
            # No mask (모든 시점을 볼 수 있음)
            mask = np.ones((n_timestep, n_timestep))
        
        return mask.astype(np.float32)
    
    
    @staticmethod
    def create_transformer_dataset(data, target, n_timestep, future_day, 
                                   stride=1, output_type='future_target'):
        """
        Transformer 학습을 위한 전체 데이터셋 생성 (원스톱 함수)
        
        Args:
            data: 원본 특징 데이터
            target: 타겟 값 (수익률 등)
            n_timestep: 시계열 길이
            future_day: 예측할 미래 일수
            stride: 윈도우 이동 간격
            output_type: 'many_to_many', 'many_to_one', 'future_target'
        
        Returns:
            input_data, output_data
            
        사용 예시:
            X, y = CreateInputOutput.create_transformer_dataset(
                data=df_scaled, 
                target=ratio,
                n_timestep=60,
                future_day=20,
                output_type='future_target'
            )
        """
        # 입력 데이터 생성
        input_data = CreateInputOutput.transformer_input_data(
            data, n_timestep, future_day, stride
        )
        
        # 출력 데이터 생성
        if output_type == 'future_target':
            output_data = CreateInputOutput.transformer_future_target(
                target, n_timestep, future_day, stride
            )
        elif output_type in ['many_to_many', 'many_to_one']:
            output_data = CreateInputOutput.transformer_output_sequence(
                target, n_timestep, future_day, stride, output_type
            )
        else:
            raise ValueError("output_type must be 'many_to_many', 'many_to_one', or 'future_target'")
        
        return input_data, output_data



class CreateTarget(object):  # target 생성
    @staticmethod
    def trend_ratio(price, future_day, alpha=100):  # 수익률
        if type(price) != np.ndarray:
            price = np.array(price)
        target = (price[future_day:] - price[0:-future_day]) / price[0:-future_day] * alpha
        return target

    @staticmethod
    def trend_ratio_avg(price, future_day, alpha=100):  # 평균 수익률
        if type(price) != np.ndarray:
            price = np.array(price)

        target_li = list()
        future_days = list(range(1, future_day + 1))

        target_li = [(price[future_day:] - price[0:-future_day]) / price[0:-future_day] * alpha for future_day in future_days]
        '''
        for future_day in future_days:
            target_li.append((price[future_day:] - price[0:-future_day]) / price[0:-future_day] * alpha)
        '''

        length = len(target_li[-1])

        for i in range(len(target_li)):
            target_li[i] = target_li[i][:length]

        target = np.vstack((target_li))
        target = (np.sum(target, axis=0) / target.shape[0])
        return target

    @staticmethod
    def log_ratio(price, future_day):
        if type(price) != np.ndarray:
            price = np.array(price)

        log_price = np.log(price)
        target = log_price[future_day:] - log_price[0:-future_day]

        return target

    @staticmethod
    def class_rise_fall(price, future_day):  # 상승/하락
        if type(price) != np.ndarray:
            price = np.array(price)
        target = list()
        diff = price[future_day:] - price[0:-future_day]

        target = [1 if i >= 0 else 0 for i in diff]
        '''
        for i in diff:
            if i >= 0:
                target.append(1)
            else:
                target.append(0)
        '''

        return np.array(target)


class InputNormalization():
    def std_scaler(train_X, test_X):
        scaler_train = StandardScaler()
        scaler_test = StandardScaler()
        # train_X_scaled = scaler_train.fit_transform(train_X) #input data 정규화
        # test_X_scaled = scaler_train.transform(test_X)
        test_X_scaled = scaler_test.fit_transform(test_X)
        return train_X_scaled, test_X_scaled

    def min_max_scaler(train_X, test_X):
        scaler = MinMaxScaler()
        train_X_scaled = scaler.fit_transform(train_X)  # input data 정규화
        test_X_scaled = scaler.transform(test_X)
        return train_X_scaled, test_X_scaled



class TimeSeriesNormalization():
    @staticmethod
    def std_scaler(data):
        #data = StandardScaler().fit_transform(data.reshape(len(data), -1)).reshape(data.shape)
        #data = data.tolist()
        #input_list = data

        input_list = list()
        print("data = ",len(data))

        for i in range(len(data)):
            input_list.append(StandardScaler().fit_transform(data[i]))

        return np.array(input_list)

    @staticmethod
    def batch_std_scaler(data, batch_size):
        size = int(len(data) / batch_size) + 1

        for i in range(size):
            if i == size:
                last_idx = len(data)
            else:
                last_idx = batch_size * (i + 1)

            data[batch_size * i:last_idx] = StandardScaler().fit_transform(data[batch_size * i:last_idx])

        return data


def remove_columns(data, nan_thresh):
    thresh = int(len(data) * nan_thresh / 100)
    col_list = list(data.columns)
    matching1 = [s for s in col_list if "date" in s]
    matching2 = [s for s in col_list if "item_code" in s]
    if 'target' in col_list:
        rm_col = matching1 + matching2 + ['target']

    rm_col = matching1 + matching2

    data = data.drop(rm_col, axis=1)
    data = data.dropna(axis=1, how='any', thresh=thresh)
    return data


class train_test_split(object):
    @staticmethod
    def sequence_time_split(data, ratio=80):
        i = round(len(data) * ratio / 100)
        return data[:i], data[i:]


def next_batch(X_data, y_data, batch_size):
    idx = np.random.choice(len(y_data), batch_size, replace=False)
    return X_data[idx], y_data[idx]


def extract_last_output(output):  # SLTM의 마지막 출력값 추출
    last_output = np.array(output)[0:, -1].reshape(-1)
    return last_output

def ratio_to_price(ratio, price, target_alpha = 100): #복원

    restored_price = (ratio / target_alpha + 1) * price
    # self.result_price = pd.DataFrame({'price_true' : self.restored_real_price, 'price_predict' : self.restored_pred_price})
    return restored_price


class Evaluate(object):
    def __init__(self, prediction, true):
        self.prediction = prediction
        self.true = true

    def MSE(self):
        result = (sum((self.prediction - self.true) ** 2) / len(self.true))

        return round(result, 4)

    def MAPE(self, predict_price, true_price):
        # true_price[np.where(true_price == 0)] = 1
        result = (sum(abs((true_price.values - predict_price.values) / true_price.values))) * 100 / len(true_price)
        return round(result, 4)


    def compute_rise_fall(self):
        self.pred_rise_fall = list()
        self.output_rise_fall = list()

        self.pred_rise_fall = [0 if self.prediction[i] < 0 else 1 for i in range(len(self.prediction))]
        self.output_rise_fall = [0 if self.true[i] < 0 else 1 for i in range(len(self.true))]

        return self.pred_rise_fall, self.output_rise_fall

    def precision_recall(self):
        self.confus_mat = confusion_matrix(self.pred_rise_fall, self.output_rise_fall)

        if self.confus_mat.shape == (1, 1):
            self.accu, self.precision_fall, self.recall_fall, self.precision_rise, self.recall_rise = 0, 0, 0, 0, 0

        self.accu = np.around(((self.confus_mat[0][0] + self.confus_mat[1][1]) / sum(sum(self.confus_mat))), decimals=3)

        self.precision_fall = round(self.confus_mat[0][0] / sum(self.confus_mat[0]), 3)
        self.recall_fall = round(self.confus_mat[0][0] / sum(self.confus_mat[:, 0]), 3)

        self.precision_rise = round(self.confus_mat[1][1] / sum(self.confus_mat[1]), 3)
        self.recall_rise = round(self.confus_mat[1][1] / sum(self.confus_mat[:, 1]), 3)

        return self.confus_mat, self.accu, self.precision_fall, self.recall_fall, self.precision_rise, self.recall_rise


class ExperimentResult:
    def __init__(self, test_predict, test_output, n_timestep, future_day):
        self.test_predict = test_predict
        self.test_output = test_output

        self.n_timestep = n_timestep
        self.future_day = future_day

    def last_output(self):  # LSTM의 마지막 출력값 추출
        self.last_test_pred = extract_last_output(self.test_predict)
        self.last_test_output = extract_last_output(self.test_output)

        #self.result_ratio = pd.DataFrame({'ratio_true' : self.last_output, 'ratio_predict' : self.last_pred})
        #return self.result_ratio
    def convert_price(self,test_price,target_alpha=100):
        self.test_predict_price = ratio_to_price(self.last_test_pred, test_price[:-self.future_day], target_alpha = target_alpha)
        # self.test_output_price = ratio_to_price(self.last_test_output, test_price, target_alpha = target_alpha)
        self.test_output_price = test_price


    def table(self, date):
        print("date shape : ", date.shape)
        self.date = date

        self.result_table = pd.DataFrame({"date":date.reset_index(drop=True),
                                          "ratio_r":self.last_test_output[self.future_day:],
                                          "ratio_p":self.last_test_pred,
                                          "price_r":self.test_output_price[self.future_day:].reset_index(drop=True),
                                          "price_p":self.test_predict_price.reset_index(drop=True)
                                        })
        return self.result_table


    def evaluation(self): #평가 지표 생성
        Ptd_Pt = self.test_output_price.values[self.future_day:] - self.test_output_price.values[:-self.future_day]
        Ptds_Pt = self.test_predict_price.values - self.test_output_price.values[:-self.future_day]
        S = np.where(np.sign(Ptd_Pt*Ptds_Pt)==0, 1, np.sign(Ptd_Pt*Ptds_Pt))
        self.PDA = len(np.where(S > 0)[0]) / len(self.date)

        T = (self.test_predict_price.values - self.test_output_price.values[self.future_day:])/\
                (self.test_output_price.values[self.future_day:] - self.test_output_price.values[:-self.future_day])

        self.PRA_M = np.mean(T)
        self.PRA_S = np.std(T)

        self.PRCR30 = len(np.where(abs(T) < 0.3)[0])/len(self.date)
        self.PRCR50 = len(np.where(abs(T) < 0.5)[0])/len(self.date)
        self.PRCR70 = len(np.where(abs(T) < 0.7)[0])/len(self.date)

        self.test_eval = Evaluate(self.last_test_pred, self.last_test_output[self.future_day:])

        self.test_MSE = self.test_eval.MSE()

        self.test_MAPE = self.test_eval.MAPE(self.test_predict_price, self.test_output_price[self.future_day:])
        self.test_eval.compute_rise_fall()

        self.test_confus_mat, self.test_accu, self.test_precision_fall, self.test_recall_fall, self.test_precision_rise, self.test_recall_rise = self.test_eval.precision_recall()

    def save_result(self, model_name, item_name, input_size, time_interval, batch_size, target_type, window_size, learning_time):
        self.info = f"{item_name}_{datetime.now().strftime('%A %d %B %Y %Hh %Mm %Ss')}"

        self.learning_time = learning_time

        print(f'info : {self.info}')
        file_name = f'{self.info}.xlsx'

        self.file_path = f'./experiment/result_excel/{file_name}'
        self.result_table.to_excel(self.file_path, index=False)

        wb = load_workbook(self.file_path, data_only=True)
        sheet1 = wb.active

        sheet1.cell(1, 7, '실험환경')
        sheet1.cell(2, 7, 'Model')
        sheet1.cell(3, 7, 'Item')
        sheet1.cell(4, 7, 'Target')
        sheet1.cell(5, 7, 'Timestep')
        sheet1.cell(6, 7, 'Interval')
        sheet1.cell(7, 7, 'Window size')
        sheet1.cell(8, 7, 'Prediction day')
        sheet1.cell(9, 7, 'batch size')

        sheet1.cell(2, 8, model_name)
        sheet1.cell(3, 8, item_name)
        sheet1.cell(4, 8, target_type)
        sheet1.cell(5, 8, self.n_timestep)
        sheet1.cell(6, 8, time_interval)
        sheet1.cell(7, 8, window_size)
        sheet1.cell(8, 8, self.future_day)
        sheet1.cell(9, 8, batch_size)

        sheet1.cell(1, 10, '실험결과')
        sheet1.cell(2, 10, 'Test Set')
        sheet1.cell(2, 11, 'TRUE')
        sheet1.cell(3, 10, 'PREDICTION')
        sheet1.cell(3, 11, '0')
        sheet1.cell(3, 12, '1')
        sheet1.cell(4, 10, '0')
        sheet1.cell(5, 10, '1')

        sheet1.cell(4, 11, self.test_confus_mat[0][0])
        sheet1.cell(4, 12, self.test_confus_mat[0][1])
        sheet1.cell(5, 11, self.test_confus_mat[1][0])
        sheet1.cell(5, 12, self.test_confus_mat[1][1])

        sheet1.cell(4, 13, 'Rise')
        sheet1.cell(5, 13, 'Fall')
        sheet1.cell(3, 15, 'Precision')
        sheet1.cell(3, 14, 'Recall')

        sheet1.cell(4, 14, self.test_precision_rise)
        sheet1.cell(4, 15, self.test_recall_rise)

        sheet1.cell(5, 14, self.test_precision_fall)
        sheet1.cell(5, 15, self.test_recall_fall)

        sheet1.cell(3, 18, 'Test')

        sheet1.cell(4, 17, 'MAPE')
        sheet1.cell(5, 17, 'MSE')
        sheet1.cell(6, 17, 'Accuracy')
        sheet1.cell(7, 17, 'PDA')
        sheet1.cell(8, 17, 'Learning_time')
        sheet1.cell(4, 19, 'PRA.M')
        sheet1.cell(5, 19, 'PRA.S')
        sheet1.cell(6, 19, 'PRCR(30%)')
        sheet1.cell(7, 19, 'PRCR(50%)')
        sheet1.cell(8, 19, 'PRCR(70%)')

        print('MAPE:', self.test_MAPE)
        sheet1.cell(4, 18, self.test_MAPE)
        sheet1.cell(5, 18, self.test_MSE)
        sheet1.cell(6, 18, self.test_accu)
        sheet1.cell(7, 18, self.PDA)
        sheet1.cell(8, 18, self.learning_time)

        sheet1.cell(4, 20, self.PRA_M)
        sheet1.cell(5, 20, self.PRA_S)
        sheet1.cell(6, 20, self.PRCR30)
        sheet1.cell(7, 20, self.PRCR50)
        sheet1.cell(8, 20, self.PRCR70)

        wb.save(self.file_path)
        return 'Succeded. Save'

    # 예측결과 DB 저장
    def upload_result(self, model_name, item_name, window_size, time_interval, prediction_day, batch_size):
        addr = "165.246.196.138"
        host = "vf-user"
        pwd = "!@value1419"
        schema = "financedb"

        # DB connection
        conn = ResultUpload_module(addr, host, pwd, schema)

        # 예측파라메터(prediction_params) 조회(없으면 생성)
        conn.creation_prediction_params(model_name, self.n_timestep, time_interval, window_size, prediction_day,
                                        batch_size)

        # 예측결과(prediction_result) 저장
        conn.creation_prediction_result(item_name, self.test_MAPE, self.test_MSE, self.test_accu,
                                        self.PDA, self.PRA_M, self.PRA_S, self.PRCR30, self.PRCR50, self.PRCR70, self.learning_time, "")

        # 예측결과상세(prediction_result_detail) 저장
        conn.creation_prediction_result_detail(self.result_table)

        return 'Succeded. Save'

    def save_visualization(self):
        print(f'MSE : {self.test_MSE},   Accuracy : {self.test_accu}')
        figsize = (15, 3)
        fig, ax = plt.subplots(1, 2, figsize=figsize)
        ax[0].set_title("RATIO")
        ax[0].set_xlabel("test-day")
        ax[0].set_ylabel("ratio")

        ax[1].set_title("PRICE")
        ax[1].set_xlabel("test-day")
        ax[1].set_ylabel("price")

        ax[0].plot(self.last_test_output, label="true")
        ax[0].plot(self.last_test_pred, label="prediction")
        ax[0].legend(loc='upper right')

        ax[1].plot(self.test_output_price.reset_index(drop=True), label="true")
        ax[1].plot(self.test_predict_price.reset_index(drop=True), label="prediction")

        ax[1].legend()

        plt.show()
        filepath = f'./experiment/result_pyplot/{self.info}.png'
        fig.savefig(filepath)

    def save_model_architecture(self, model, model_name):
        keras.utils.plot_model(model, f'./experiment/model_architecture/{model_name}_model_with_shape_info.png', show_shapes=True)

    def save_model(self, model):
        file_path = f'./experiment/models/{self.info}'
        tf.saved_model.save(model, file_path)


def predict_batch_test(model, test_input, batch_size):
    int(model.output.shape[1])
    int(model.output.shape[2])
    pred = np.zeros((test_input.shape[0], test_input.shape[1], int(model.output.shape[2])), dtype=np.float32)

    for i in range(int(len(test_input) / batch_size)):
        result = model(test_input[i * batch_size:(i + 1) * batch_size], training=False)
        # pred += list(result)
        pred[i * batch_size:(i + 1) * batch_size] = result

    if len(test_input) % batch_size != 0:
        result = model(test_input[(i + 1) * batch_size:], training=False)

        pred[(i + 1) * batch_size:] = result
    print("Numpy Converting...")
    print("[Success] Numpy conversion")
    return pred


def create_learning_data(item_name, n_timestep, future_day, time_interval, price_col_name, target_type, fill_na, channel=True):
    dataframe = DataLoad.fndata_dataframe(item_name, merge=True, fill_na=fill_na)
    train_data, test_data = train_test_split.sequence_time_split(dataframe, ratio=80)  # CLNN1

    # 학습시 불필요한 변수 제거
    train_data = remove_columns(train_data)
    test_data = remove_columns(test_data)

    train_scaled, test_scaled = InputNormalization.std_scaler(train_data, test_data)

    # 모델 Input data 생성
    train_input = CreateInputOutput.input_2d_data(train_scaled, n_timestep, future_day, time_interval, channel=channel)
    test_input = CreateInputOutput.input_2d_data(test_scaled, n_timestep, future_day, time_interval, channel=channel)

    # 종목의 가격 추출
    train_price = train_data[price_col_name]
    test_price = test_data[price_col_name]

    # 수익률 계산
    if target_type == 'trend_ratio':
        train_target = CreateTarget.trend_ratio(train_price, future_day)
        test_target = CreateTarget.trend_ratio(test_price, future_day)
    elif target_type == 'log_ratio':
        train_target = CreateTarget.log_ratio(train_price, future_day)
        test_target = CreateTarget.log_ratio(test_price, future_day)

    # 모델 output 생성
    train_output = CreateInputOutput.many_to_many_output(train_target, n_timestep)
    test_output = CreateInputOutput.many_to_many_output(test_target, n_timestep)

    print(f'train input size : {train_input.shape},      train output size : {train_output.shape}')
    print(f'test input size : {test_input.shape},     test output size : {test_output.shape}')

    return train_input, train_output, test_input, test_output, test_price